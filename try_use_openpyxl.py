from openpyxl import Workbook#不需要设置样式，只需要导入这个类
#from openpyxl.styles import Font,PatternFill#如果需要设置字体、背景颜色之类需要导入这些类
from openpyxl.styles import Font ,PatternFill,Border,Side#要设置粗体

#一个工作簿（Workbook）可以有多张工作表（Worksheet）

#先实例化一个对象
wb = Workbook()#Workbook对象代表工作簿 新创建的工作簿默认会包含一个工作表
ws = wb.active#这个属性，得到工作簿当前活动的工作表，即用户打开Excel时默认选中的工作表
#因为active属性返回的数据类型是Worksheet对象，可以赋值给一个ws变量,之后就可以通过这个对象向对应的工作表中读写数据了
#比如可对ws的title属性赋值，来修改工作表的标签名称
ws.title = '豆瓣电影top250'
#接下来可以在工作表的表头，即第一行写入数据
headers = ['排名','电影名称','评分','年份']
#先创建一个储存表头的列表headers，然后通过for循环来依次写入每个列名
for i in range(len(headers)):
    header = headers[i]
    #要写入数据，我们需要能表示写入位置，即第几行第几列，但excel的行号和列号是从1开始的，这个和python里索引从0开始有区别
    col_num = i + 1
    #所以让表示列数的col_num变量在i的基础上加1
    cell = ws.cell(row=1, column=col_num)
    #然后调用ws的cell方法，通过指定行号row和列号column，得到对应位置的Cell对象，也就是表示单元格的对象
    cell.value = header
    #然后通过设置cell对象的value属性，给单元格写入数据值,也就是列名

    cell.font = Font(bold=True) #把bold，也就是粗体参数设置为True，然后把这个实例赋值给表示表头的每个单元格的font属性
    #除此之外，Font类的构造函数提供多种参数来控制字体样式
    cell.font = Font(bold=True,name = '黑体',size = 22,color = "FF0000",underline='double')
    #执行代码前要先关闭Excel文件，否则执行时可能出现权限异常

    #除了字体，还可以设置背景颜色，用styles子模块的PatternFill类
    #同样，要实例化这个类,然后赋值给单元格的fill属性
    #cell.fill = PatternFill(patternType='solid',fgColor='FDF5E6')
    # PatternFill表示填充类型，solid表示纯色，参数fgcolor用来指定填充颜色
    #还可以设置bgcolor参数，用来指定背景色

    cell.fill = PatternFill(patternType='darkUp', fgColor='FDF5E6',bgColor='0000FF')
    #接下来把填充类型改为darkup（上斜线），并设置bgcolor让斜线颜色为蓝色
    #打开表格可以看到单元格被增加了上斜线条纹填充，颜色也对应传入的十六进制

    #还可以设置单元格的边框样式，包括边框的线条类型和颜色
    #这就需要送styles子模块导入Border和Side类
    #Side类用于定义边框单边的样式属性，包括线条类型、颜色
    #Border类用于组合四个Side对象，分别应用于单元格的上下左右边框
    thick_side = Side(style = "thick",color = "00FF00")
    #在Side构造函数里，参数Style用于设置线条样式，thick表示粗线，然后可以把这个单边的样式属性应用到单元格的border属性上
    cell.border = Border(top = thick_side, right = thick_side, bottom = thick_side, left = thick_side)
    #那么再实例化一个Border对象，上、下、左、右边框的样式，都可以通过Top、bottom、left、right参数分别设置
    #如果想进一步定制边框样式，style的可选值还有dashDot（点划线）、hair（极细线）等

#接下来把电影数据写入到工作表中
#和表头类似，也是要用循环对每行中的每个单元格依次写入数据
#首先遍历电影数据列表movies,把包含各个电影信息的列表提取出来
movies = [
    [1,"肖申克的救赎",9.7,1994],
    [2,"霸王别姬",9.6,1993],
]
# for i in range(len(movies)):
#     movie = movies[i]
#     row = i+2#因为每个电影占一行，所以把行号row赋值为从2开始依次递增的数字
#     #在获得每部电影对应的相关信息列表以及写入的行数后，我们进一步循环电影信息列表
#     #把对应到各个单元格的数据提取出来，赋值给value变量
#     for j in range(len(movie)):
#         value = movie[j]
#         col = j+1#然后我们还需要得到每个单元格的列数，这里就是循环变量j+1
#         ws.cell(row = row,column=col,value = value)
#         #最后调用工作表的cell方法，通过对value参数进行设置，直接填入数据

#但是还有比上面这个for循环更简单的写法：
#可以用python自带的enumerate函数
#enumerate（可遍历对象）的用法是这样的：传入一个列表/字符串/元组等可以遍历的对象作为参数
#它就会帮我们同时获取索引和元素值
#但是它更进阶在于，它可以自定义起始索引，只需要再传入一个参数
for row,movie in enumerate(movies,2):
    for col,value in enumerate(movie,1):
        ws.cell(row=row,column=col,value = value)

#如果觉得用嵌套循环给每一个单元格依次设置value很麻烦
#也可以用工作表对象的append方法，直接在工作表的最后一行下添加一行新数据
ws.append([3,"阿甘正传",9.5,1994])
ws.append([4,"美丽人生",9.5,1997])
#我们给append方法传递的参数可以是列表、元组和字典
#列表和元组作为参数时，数据会按照顺序进行填入
#而传递字典时，键应该是比如"A"、"B"、"C"之类的列字母，表示数据对应要写入的列
ws.append({'A':5, 'B':"千与千寻", 'C':9.3, 'D':2001})

#写入数据后，我们可以单独修改指定单元格的值
#在Excel中，单元格位置可以用“A1表示法”，其中列用字母表示，比如A、B、C，行用数字表示，比如1、2、3
#每个单元格都有自己的地址，比如第一行第一列的单元格的地址是“A1”
#如果想修改B3这个值为霸王别姬的单元格，可以先通过地址得到这个单元格对象，然后修改对象对应的属性
#要获得单元格对象，我们在工作表对象ws后加上索引操作符[],里面放入地址
cell_b3 = ws['B3']
cell_b3.value = '霸王霸王'
cell_b3.font = Font(color = "00FF00")
cell_b3_size = Side(style = "thick",color = "00FFFF")
cell_b3.border = Border(left = cell_b3_size, right = cell_b3_size)



wb.save('豆瓣电影Top250.xlsx')
#数据写好后就可以用工作簿对象wb的save方法来保存excel文件了



